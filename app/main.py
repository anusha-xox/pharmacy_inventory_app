from datetime import datetime, timedelta, date
from pathlib import Path
import os
import sqlite3
import smtplib
from email.message import EmailMessage
from typing import Optional
import zipfile
import io
import shutil

from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, Field

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "pharmacy.db"

app = FastAPI(title="Hospital Pharmacy Inventory & Billing System")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
templates = Jinja2Templates(directory=BASE_DIR / "templates")


def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def d(row):
    return dict(row) if row else None


def init_db():
    conn = db()
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        user_id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL DEFAULT 'demo123',
        role TEXT NOT NULL CHECK(role IN ('admin','desk')),
        is_active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS medicines (
        medicine_id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        strength TEXT,
        category TEXT,
        manufacturer TEXT,
        reorder_level INTEGER NOT NULL DEFAULT 10,
        is_active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS medicine_batches (
        batch_id INTEGER PRIMARY KEY AUTOINCREMENT,
        medicine_id INTEGER NOT NULL,
        batch_no TEXT NOT NULL,
        quantity_available INTEGER NOT NULL CHECK(quantity_available >= 0),
        date_of_adding TEXT NOT NULL,
        expiry_date TEXT NOT NULL,
        unit_price REAL NOT NULL DEFAULT 0,
        supplier TEXT,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (medicine_id) REFERENCES medicines(medicine_id)
    );
    CREATE TABLE IF NOT EXISTS bills (
        bill_id INTEGER PRIMARY KEY AUTOINCREMENT,
        bill_no TEXT UNIQUE NOT NULL,
        customer_name TEXT,
        customer_email TEXT,
        subtotal REAL NOT NULL DEFAULT 0,
        discount REAL NOT NULL DEFAULT 0,
        total_amount REAL NOT NULL,
        payment_method TEXT NOT NULL DEFAULT 'Cash',
        created_by INTEGER,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (created_by) REFERENCES users(user_id)
    );
    CREATE TABLE IF NOT EXISTS bill_items (
        bill_item_id INTEGER PRIMARY KEY AUTOINCREMENT,
        bill_id INTEGER NOT NULL,
        medicine_id INTEGER NOT NULL,
        batch_id INTEGER NOT NULL,
        quantity INTEGER NOT NULL,
        unit_price REAL NOT NULL,
        line_total REAL NOT NULL,
        FOREIGN KEY (bill_id) REFERENCES bills(bill_id),
        FOREIGN KEY (medicine_id) REFERENCES medicines(medicine_id),
        FOREIGN KEY (batch_id) REFERENCES medicine_batches(batch_id)
    );
    CREATE TABLE IF NOT EXISTS stock_movements (
        movement_id INTEGER PRIMARY KEY AUTOINCREMENT,
        medicine_id INTEGER NOT NULL,
        batch_id INTEGER,
        movement_type TEXT NOT NULL,
        quantity_change INTEGER NOT NULL,
        reference_id INTEGER,
        notes TEXT,
        created_by INTEGER,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (medicine_id) REFERENCES medicines(medicine_id),
        FOREIGN KEY (batch_id) REFERENCES medicine_batches(batch_id),
        FOREIGN KEY (created_by) REFERENCES users(user_id)
    );
    """)
    if conn.execute("SELECT COUNT(*) c FROM users").fetchone()["c"] == 0:
        conn.executemany("INSERT INTO users(name,username,password,role) VALUES(?,?,?,?)", [
            ("Admin", "admin", "demo123", "admin"), ("Desk User", "desk", "demo123", "desk")])
        meds = [
            ("Paracetamol", "500mg", "Tablet", "Calpol", 20),
            ("Amoxicillin", "250mg", "Capsule", "Cipla", 15),
            ("Cetirizine", "10mg", "Tablet", "Dr Reddy", 20),
            ("Ibuprofen", "400mg", "Tablet", "Abbott", 20),
            ("Azithromycin", "500mg", "Tablet", "Sun Pharma", 10),
        ]
        conn.executemany("INSERT INTO medicines(name,strength,category,manufacturer,reorder_level) VALUES(?,?,?,?,?)", meds)
        # Sample batches with date_of_adding, expiry_date, unit_price
        batches = [
            (1, "2024-05-01", "2026-08-31", 100, 15, "Medlife Pharma"),
            (1, "2024-06-15", "2026-12-31", 50, 15, "Medlife Pharma"),
            (2, "2024-05-10", "2026-10-15", 80, 32, "HealthRx"),
            (3, "2024-05-20", "2026-11-30", 120, 8, "MediSupply"),
            (4, "2024-06-01", "2026-09-05", 60, 20, "HealthRx"),
            (5, "2024-05-25", "2026-06-20", 15, 45, "MediSupply")
        ]
        for med_id, date_add, expiry, qty, unit_price, supplier in batches:
            # Get medicine name for batch_no generation
            med_name = conn.execute("SELECT name FROM medicines WHERE medicine_id=?", (med_id,)).fetchone()["name"]
            batch_no = f"{date_add}_{med_name}_{expiry}"
            conn.execute(
                "INSERT INTO medicine_batches(medicine_id,batch_no,date_of_adding,expiry_date,quantity_available,unit_price,supplier) VALUES(?,?,?,?,?,?,?)",
                (med_id, batch_no, date_add, expiry, qty, unit_price, supplier)
            )
            bid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            conn.execute("INSERT INTO stock_movements(medicine_id,batch_id,movement_type,quantity_change,notes,created_by) VALUES(?,?, 'stock_added', ?, 'Initial stock seed', 1)", (med_id, bid, qty))
    # Add customer_phone column if it doesn't exist
    try:
        conn.execute("ALTER TABLE bills ADD COLUMN customer_phone TEXT")
        conn.commit()
    except sqlite3.OperationalError:
        pass  # Column already exists
    
    conn.commit(); conn.close()


class LoginIn(BaseModel):
    username: str
    password: str
    role: str
class UserIn(BaseModel):
    name: str; username: str; password: str = "demo123"; role: str
class MedicineIn(BaseModel):
    name: str; strength: Optional[str]=""; category: Optional[str]=""; manufacturer: Optional[str]=""; reorder_level: int = Field(default=10, ge=0)
class MedicineUpdate(MedicineIn):
    is_active: int = 1
class BatchIn(BaseModel):
    medicine_id: int; date_of_adding: str; expiry_date: str; quantity_available: int = Field(ge=0); unit_price: float = Field(ge=0); supplier: Optional[str]=""; created_by: int = 1
class BillItemIn(BaseModel):
    medicine_id: int; quantity: int = Field(gt=0); batch_id: Optional[int] = None
class BillIn(BaseModel):
    customer_name: Optional[str]="Walk-in Customer"; customer_email: Optional[str]=""; customer_phone: Optional[str]=""; created_by: int = 2; items: list[BillItemIn]; discount: float = 0; payment_method: str = "Cash"
class EmailBillIn(BaseModel):
    to_email: str
    smtp_username: Optional[str] = None
    smtp_password: Optional[str] = None
    smtp_server: Optional[str] = None
    smtp_port: Optional[int] = None

@app.on_event("startup")
def startup(): init_db()

@app.get("/", response_class=HTMLResponse)
def home(request: Request): return templates.TemplateResponse("index.html", {"request": request})

@app.post("/api/login")
def login(payload: LoginIn):
    conn = db(); row = conn.execute("SELECT user_id,name,username,role FROM users WHERE username=? AND password=? AND role=? AND is_active=1", (payload.username, payload.password, payload.role)).fetchone(); conn.close()
    if not row: raise HTTPException(401, "Invalid login. Demo: admin/demo123 or desk/demo123")
    return d(row)

@app.get("/api/users")
def users():
    conn=db(); rows=conn.execute("SELECT user_id,name,username,role,is_active,created_at FROM users ORDER BY role,name").fetchall(); conn.close(); return [d(r) for r in rows]
@app.post("/api/users")
def add_user(u:UserIn):
    conn=db()
    try:
        cur=conn.execute("INSERT INTO users(name,username,password,role) VALUES(?,?,?,?)", (u.name,u.username,u.password,u.role)); conn.commit(); return {"user_id": cur.lastrowid}
    except sqlite3.IntegrityError: raise HTTPException(400,"Username already exists")
    finally: conn.close()
@app.patch("/api/users/{user_id}/toggle")
def toggle_user(user_id:int):
    conn=db(); conn.execute("UPDATE users SET is_active = CASE WHEN is_active=1 THEN 0 ELSE 1 END WHERE user_id=?",(user_id,)); conn.commit(); conn.close(); return {"ok":True}

@app.get("/api/medicines")
def medicines(q: str = "", include_inactive: int = 0):
    conn=db(); active_clause = "" if include_inactive else "AND m.is_active=1"
    rows=conn.execute(f"""
    SELECT m.*, COALESCE(SUM(CASE WHEN b.expiry_date >= date('now') THEN b.quantity_available ELSE 0 END),0) total_stock,
    MIN(CASE WHEN b.quantity_available>0 AND b.expiry_date >= date('now') THEN b.expiry_date END) nearest_expiry
    FROM medicines m LEFT JOIN medicine_batches b ON b.medicine_id=m.medicine_id
    WHERE (m.name || ' ' || IFNULL(m.strength,'') || ' ' || IFNULL(m.manufacturer,'')) LIKE ? {active_clause}
    GROUP BY m.medicine_id ORDER BY m.is_active DESC, m.name""", (f"%{q}%",)).fetchall(); conn.close(); return [d(r) for r in rows]
@app.post("/api/medicines")
def add_medicine(m:MedicineIn):
    conn=db(); cur=conn.execute("INSERT INTO medicines(name,strength,category,manufacturer,reorder_level) VALUES(?,?,?,?,?)", (m.name,m.strength,m.category,m.manufacturer,m.reorder_level)); conn.commit(); conn.close(); return {"medicine_id":cur.lastrowid}
@app.put("/api/medicines/{medicine_id}")
def update_medicine(medicine_id:int,m:MedicineUpdate):
    conn=db(); conn.execute("UPDATE medicines SET name=?,strength=?,category=?,manufacturer=?,reorder_level=?,is_active=? WHERE medicine_id=?", (m.name,m.strength,m.category,m.manufacturer,m.reorder_level,m.is_active,medicine_id)); conn.commit(); conn.close(); return {"ok":True}
@app.patch("/api/medicines/{medicine_id}/toggle")
def toggle_medicine(medicine_id:int):
    conn=db(); conn.execute("UPDATE medicines SET is_active=CASE WHEN is_active=1 THEN 0 ELSE 1 END WHERE medicine_id=?",(medicine_id,)); conn.commit(); conn.close(); return {"ok":True}

@app.get("/api/medicines/{medicine_id}/batches")
def batches(medicine_id:int, active_only:int=1):
    conn=db(); clause="AND quantity_available>0 AND expiry_date>=date('now')" if active_only else ""
    rows=conn.execute(f"SELECT * FROM medicine_batches WHERE medicine_id=? {clause} ORDER BY expiry_date ASC", (medicine_id,)).fetchall(); conn.close(); return [d(r) for r in rows]
@app.post("/api/batches")
def add_batch(b:BatchIn):
    conn=db()
    # Get medicine name to generate batch_no
    med = conn.execute("SELECT name FROM medicines WHERE medicine_id=?", (b.medicine_id,)).fetchone()
    if not med:
        conn.close()
        raise HTTPException(404, "Medicine not found")
    
    # Generate batch_no: date_of_adding_medicine + medicine name + expiry date
    batch_no = f"{b.date_of_adding}_{med['name']}_{b.expiry_date}"
    
    cur=conn.execute("INSERT INTO medicine_batches(medicine_id,batch_no,date_of_adding,expiry_date,quantity_available,unit_price,supplier) VALUES(?,?,?,?,?,?,?)", (b.medicine_id,batch_no,b.date_of_adding,b.expiry_date,b.quantity_available,b.unit_price,b.supplier))
    bid=cur.lastrowid
    conn.execute("INSERT INTO stock_movements(medicine_id,batch_id,movement_type,quantity_change,notes,created_by) VALUES(?,?, 'stock_added', ?, 'Batch stock added', ?)", (b.medicine_id,bid,b.quantity_available,b.created_by))
    conn.commit()
    conn.close()
    return {"batch_id":bid, "batch_no": batch_no}

@app.put("/api/batches/{batch_id}")
def update_batch(batch_id:int, b:BatchIn):
    conn=db()
    # Get existing batch to calculate quantity change
    existing = conn.execute("SELECT * FROM medicine_batches WHERE batch_id=?", (batch_id,)).fetchone()
    if not existing:
        conn.close()
        raise HTTPException(404, "Batch not found")
    
    # Get medicine name to regenerate batch_no
    med = conn.execute("SELECT name FROM medicines WHERE medicine_id=?", (b.medicine_id,)).fetchone()
    if not med:
        conn.close()
        raise HTTPException(404, "Medicine not found")
    
    batch_no = f"{b.date_of_adding}_{med['name']}_{b.expiry_date}"
    
    # Update batch
    conn.execute("UPDATE medicine_batches SET medicine_id=?,batch_no=?,date_of_adding=?,expiry_date=?,quantity_available=?,unit_price=?,supplier=? WHERE batch_id=?",
                 (b.medicine_id,batch_no,b.date_of_adding,b.expiry_date,b.quantity_available,b.unit_price,b.supplier,batch_id))
    
    # Record stock movement if quantity changed
    qty_change = b.quantity_available - existing["quantity_available"]
    if qty_change != 0:
        conn.execute("INSERT INTO stock_movements(medicine_id,batch_id,movement_type,quantity_change,notes,created_by) VALUES(?,?, 'adjustment', ?, 'Batch quantity adjusted', ?)",
                     (b.medicine_id,batch_id,qty_change,b.created_by))
    
    conn.commit()
    conn.close()
    return {"ok":True, "batch_no": batch_no}

@app.get("/api/alerts")
def alerts():
    conn=db(); low=conn.execute("""SELECT m.*, COALESCE(SUM(CASE WHEN b.expiry_date>=date('now') THEN b.quantity_available ELSE 0 END),0) total_stock FROM medicines m LEFT JOIN medicine_batches b ON b.medicine_id=m.medicine_id WHERE m.is_active=1 GROUP BY m.medicine_id HAVING total_stock <= m.reorder_level ORDER BY total_stock""").fetchall()
    exp=conn.execute("""SELECT b.*,m.name,m.strength FROM medicine_batches b JOIN medicines m ON m.medicine_id=b.medicine_id WHERE b.quantity_available>0 AND b.expiry_date BETWEEN date('now') AND date('now','+90 day') ORDER BY b.expiry_date""").fetchall()
    expired=conn.execute("""SELECT b.*,m.name,m.strength FROM medicine_batches b JOIN medicines m ON m.medicine_id=b.medicine_id WHERE b.quantity_available>0 AND b.expiry_date < date('now') ORDER BY b.expiry_date""").fetchall(); conn.close(); return {"low_stock":[d(r) for r in low],"expiring_soon":[d(r) for r in exp],"expired":[d(r) for r in expired]}

@app.post("/api/bills")
def create_bill(bill:BillIn):
    if not bill.items: raise HTTPException(400,"Cart is empty")
    conn=db()
    try:
        subtotal=0.0; deductions=[]
        for item in bill.items:
            med=conn.execute("SELECT * FROM medicines WHERE medicine_id=? AND is_active=1", (item.medicine_id,)).fetchone()
            if not med: raise HTTPException(404,"Medicine not found or inactive")
            remaining=item.quantity
            if item.batch_id:
                batch_rows=conn.execute("SELECT * FROM medicine_batches WHERE batch_id=? AND medicine_id=? AND quantity_available>0 AND expiry_date>=date('now')", (item.batch_id,item.medicine_id)).fetchall()
            else:
                batch_rows=conn.execute("SELECT * FROM medicine_batches WHERE medicine_id=? AND quantity_available>0 AND expiry_date>=date('now') ORDER BY expiry_date ASC", (item.medicine_id,)).fetchall()
            for batch in batch_rows:
                if remaining<=0: break
                # Use unit_price from batch instead of medicine
                unit_price = batch["unit_price"]
                qty=min(remaining,batch["quantity_available"]); line=qty*unit_price; subtotal+=line; remaining-=qty
                deductions.append({"medicine_id":item.medicine_id,"batch_id":batch["batch_id"],"qty":qty,"unit_price":unit_price,"line_total":line})
            if remaining>0: raise HTTPException(400, f"Not enough stock for {med['name']} {med['strength'] or ''}")
        discount=max(0,min(bill.discount,subtotal)); total=subtotal-discount
        bill_no=f"INV-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        cur=conn.execute("INSERT INTO bills(bill_no,customer_name,customer_email,customer_phone,subtotal,discount,total_amount,payment_method,created_by) VALUES(?,?,?,?,?,?,?,?,?)", (bill_no,bill.customer_name,bill.customer_email,bill.customer_phone,subtotal,discount,total,bill.payment_method,bill.created_by)); bill_id=cur.lastrowid
        for x in deductions:
            conn.execute("UPDATE medicine_batches SET quantity_available=quantity_available-? WHERE batch_id=?", (x["qty"],x["batch_id"]))
            conn.execute("INSERT INTO bill_items(bill_id,medicine_id,batch_id,quantity,unit_price,line_total) VALUES(?,?,?,?,?,?)", (bill_id,x["medicine_id"],x["batch_id"],x["qty"],x["unit_price"],x["line_total"]))
            conn.execute("INSERT INTO stock_movements(medicine_id,batch_id,movement_type,quantity_change,reference_id,notes,created_by) VALUES(?,?, 'sale', ?, ?, 'Billing deduction', ?)", (x["medicine_id"],x["batch_id"],-x["qty"],bill_id,bill.created_by))
        conn.commit(); return get_bill_data(conn,bill_id)
    finally: conn.close()

def get_bill_data(conn,bill_id:int):
    bill=conn.execute("SELECT b.*,u.name created_by_name FROM bills b LEFT JOIN users u ON u.user_id=b.created_by WHERE bill_id=?", (bill_id,)).fetchone()
    if not bill: raise HTTPException(404,"Bill not found")
    items=conn.execute("""SELECT bi.*,m.name,m.strength,mb.batch_no,mb.expiry_date FROM bill_items bi JOIN medicines m ON m.medicine_id=bi.medicine_id JOIN medicine_batches mb ON mb.batch_id=bi.batch_id WHERE bi.bill_id=? ORDER BY bi.bill_item_id""", (bill_id,)).fetchall()
    out=d(bill); out["items"]=[d(i) for i in items]; 
    return out

@app.get("/api/bills")
def get_bills(q: str = ""):
    conn = db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Build search query
    search_clause = ""
    params = []
    if q:
        search_clause = """
            WHERE (b.customer_name LIKE ? OR b.customer_email LIKE ? OR b.customer_phone LIKE ? OR b.bill_no LIKE ?)
        """
        search_param = f"%{q}%"
        params = [search_param, search_param, search_param, search_param]

    bills = cur.execute(f"""
        SELECT
            b.bill_id,
            b.bill_id AS id,
            b.bill_no,
            b.customer_name,
            b.customer_email,
            b.customer_phone,
            b.subtotal,
            b.discount,
            b.total_amount,
            b.payment_method,
            b.created_at,
            u.name AS created_by_name,
            u.username AS created_by_username
        FROM bills b
        LEFT JOIN users u ON b.created_by = u.user_id
        {search_clause}
        ORDER BY datetime(b.created_at) DESC
    """, params).fetchall()

    result = []

    for bill in bills:
        items = cur.execute("""
            SELECT 
                bi.bill_item_id,
                bi.bill_item_id AS id,
                bi.medicine_id,
                bi.batch_id,
                bi.quantity,
                bi.unit_price,
                bi.line_total,
                m.name AS medicine_name,
                m.name,
                m.strength,
                mb.batch_no,
                mb.expiry_date
            FROM bill_items bi
            LEFT JOIN medicines m ON bi.medicine_id = m.medicine_id
            LEFT JOIN medicine_batches mb ON bi.batch_id = mb.batch_id
            WHERE bi.bill_id = ?
            ORDER BY bi.bill_item_id
        """, (bill["bill_id"],)).fetchall()

        result.append({
            "bill_id": bill["bill_id"],
            "id": bill["bill_id"],
            "bill_no": bill["bill_no"],
            "customer_name": bill["customer_name"] or "Walk-in Customer",
            "customer_email": bill["customer_email"] or "",
            "subtotal": bill["subtotal"] or 0,
            "discount": bill["discount"] or 0,
            "total_amount": bill["total_amount"] or 0,
            "payment_method": bill["payment_method"] or "Cash",
            "created_at": bill["created_at"],
            "created_by_name": bill["created_by_name"] or "Unknown",
            "created_by_username": bill["created_by_username"] or "",
            "items": [dict(item) for item in items]
        })

    conn.close()
    return result


@app.get("/api/bills/{bill_id}")
def bill_detail(bill_id:int):
    conn=db(); out=get_bill_data(conn,bill_id); conn.close(); return out

@app.get("/api/stock-movements")
def stock_movements():
    conn=db(); rows=conn.execute("""SELECT sm.*,m.name,m.strength,mb.batch_no,u.name created_by_name FROM stock_movements sm JOIN medicines m ON m.medicine_id=sm.medicine_id LEFT JOIN medicine_batches mb ON mb.batch_id=sm.batch_id LEFT JOIN users u ON u.user_id=sm.created_by ORDER BY sm.created_at DESC LIMIT 200""").fetchall(); conn.close(); return [d(r) for r in rows]
@app.get("/api/reports")
def reports():
    conn = db()
    summary = conn.execute("""
        SELECT COALESCE(SUM(total_amount),0) total_sales,
               COUNT(*) total_bills,
               COALESCE(AVG(total_amount),0) avg_bill
        FROM bills
        WHERE date(created_at) >= date('now','-30 day')
    """).fetchone()

    # Today graph: hourly sales from 00:00 to 23:00
    today_rows = conn.execute("""
        SELECT strftime('%H', created_at) bucket,
               SUM(total_amount) sales,
               COUNT(*) bills
        FROM bills
        WHERE date(created_at) = date('now')
        GROUP BY bucket
        ORDER BY bucket
    """).fetchall()
    today_map = {r["bucket"]: d(r) for r in today_rows}
    today = []
    for h in range(24):
        key = f"{h:02d}"
        row = today_map.get(key, {})
        today.append({
            "label": f"{h:02d}:00",
            "sales": float(row.get("sales", 0) or 0),
            "bills": int(row.get("bills", 0) or 0)
        })

    # Week graph: daily sales for the last 7 days including today
    week_rows = conn.execute("""
        SELECT date(created_at) bucket,
               SUM(total_amount) sales,
               COUNT(*) bills
        FROM bills
        WHERE date(created_at) >= date('now','-6 day')
        GROUP BY bucket
        ORDER BY bucket
    """).fetchall()
    week_map = {r["bucket"]: d(r) for r in week_rows}
    week = []
    for i in range(6, -1, -1):
        dt = date.today() - timedelta(days=i)
        key = dt.isoformat()
        row = week_map.get(key, {})
        week.append({
            "label": dt.strftime('%d %b'),
            "sales": float(row.get("sales", 0) or 0),
            "bills": int(row.get("bills", 0) or 0)
        })

    # Month graph: daily sales for the last 30 days including today
    month_rows = conn.execute("""
        SELECT date(created_at) bucket,
               SUM(total_amount) sales,
               COUNT(*) bills
        FROM bills
        WHERE date(created_at) >= date('now','-29 day')
        GROUP BY bucket
        ORDER BY bucket
    """).fetchall()
    month_map = {r["bucket"]: d(r) for r in month_rows}
    month = []
    for i in range(29, -1, -1):
        dt = date.today() - timedelta(days=i)
        key = dt.isoformat()
        row = month_map.get(key, {})
        month.append({
            "label": dt.strftime('%d %b'),
            "sales": float(row.get("sales", 0) or 0),
            "bills": int(row.get("bills", 0) or 0)
        })

    top = conn.execute("""
        SELECT m.name || ' ' || IFNULL(m.strength,'') medicine, SUM(bi.quantity) qty
        FROM bill_items bi
        JOIN medicines m ON m.medicine_id=bi.medicine_id
        JOIN bills b ON b.bill_id=bi.bill_id
        WHERE date(b.created_at)>=date('now','-30 day')
        GROUP BY m.medicine_id
        ORDER BY qty DESC
        LIMIT 5
    """).fetchall()
    conn.close()
    return {
        "summary": d(summary),
        "charts": {"today": today, "week": week, "month": month},
        "daily": week,  # backward compatible fallback for older app.js
        "top_medicines": [d(r) for r in top]
    }


def bill_email_body_text(b):
    """Plain text version for email clients that don't support HTML"""
    lines=[f"Bill No: {b['bill_no']}", f"Customer: {b.get('customer_name') or 'Walk-in Customer'}"]
    if b.get('customer_phone'):
        lines.append(f"Phone: {b['customer_phone']}")
    lines.extend([f"Date: {b['created_at']}", f"Payment: {b['payment_method']}", "", "Items:"])
    for i in b["items"]:
        lines.append(f"- {i['name']} {i.get('strength') or ''} | Batch {i['batch_no']} | Qty {i['quantity']} | ₹{i['unit_price']:.2f} | ₹{i['line_total']:.2f}")
    lines += ["", f"Subtotal: ₹{b['subtotal']:.2f}", f"Discount: ₹{b['discount']:.2f}", f"Total: ₹{b['total_amount']:.2f}", "", "Thank you."]
    return "\n".join(lines)

def bill_email_body_html(b):
    """Beautiful HTML email template"""
    items_html = ""
    for i in b["items"]:
        items_html += f"""
        <tr>
            <td style="padding: 12px; border-bottom: 1px solid #e5e7eb;">
                <strong>{i['name']} {i.get('strength') or ''}</strong><br>
                <span style="color: #6b7280; font-size: 13px;">Batch: {i['batch_no']}</span>
            </td>
            <td style="padding: 12px; border-bottom: 1px solid #e5e7eb; text-align: center;">{i['quantity']}</td>
            <td style="padding: 12px; border-bottom: 1px solid #e5e7eb; text-align: right;">₹{i['unit_price']:.2f}</td>
            <td style="padding: 12px; border-bottom: 1px solid #e5e7eb; text-align: right;"><strong>₹{i['line_total']:.2f}</strong></td>
        </tr>
        """
    
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
    </head>
    <body style="margin: 0; padding: 0; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif; background-color: #f3f4f6;">
        <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #f3f4f6; padding: 40px 20px;">
            <tr>
                <td align="center">
                    <table width="600" cellpadding="0" cellspacing="0" style="background-color: #ffffff; border-radius: 12px; box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1); overflow: hidden;">
                        <!-- Header -->
                        <tr>
                            <td style="background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%); padding: 40px 30px; text-align: center;">
                                <h1 style="margin: 0; color: #ffffff; font-size: 28px; font-weight: 600;">Hospital Pharmacy</h1>
                                <p style="margin: 10px 0 0 0; color: #dbeafe; font-size: 14px;">Pharmacy Invoice</p>
                            </td>
                        </tr>
                        
                        <!-- Bill Info -->
                        <tr>
                            <td style="padding: 30px;">
                                <table width="100%" cellpadding="0" cellspacing="0">
                                    <tr>
                                        <td style="padding-bottom: 20px;">
                                            <div style="background-color: #f0f9ff; border-left: 4px solid #3b82f6; padding: 15px; border-radius: 4px;">
                                                <table width="100%" cellpadding="0" cellspacing="0">
                                                    <tr>
                                                        <td style="padding: 5px 0;">
                                                            <span style="color: #6b7280; font-size: 13px;">Bill Number</span><br>
                                                            <strong style="color: #1f2937; font-size: 16px;">{b['bill_no']}</strong>
                                                        </td>
                                                        <td style="padding: 5px 0; text-align: right;">
                                                            <span style="color: #6b7280; font-size: 13px;">Date</span><br>
                                                            <strong style="color: #1f2937; font-size: 16px;">{b['created_at']}</strong>
                                                        </td>
                                                    </tr>
                                                    <tr>
                                                        <td style="padding: 10px 0 5px 0;">
                                                            <span style="color: #6b7280; font-size: 13px;">Customer</span><br>
                                                            <strong style="color: #1f2937; font-size: 16px;">{b.get('customer_name') or 'Walk-in Customer'}</strong>
                                                            {f'<br><span style="color: #6b7280; font-size: 13px;">Phone: {b["customer_phone"]}</span>' if b.get('customer_phone') else ''}
                                                        </td>
                                                        <td style="padding: 10px 0 5px 0; text-align: right;">
                                                            <span style="color: #6b7280; font-size: 13px;">Payment Method</span><br>
                                                            <strong style="color: #1f2937; font-size: 16px;">{b['payment_method']}</strong>
                                                        </td>
                                                    </tr>
                                                </table>
                                            </div>
                                        </td>
                                    </tr>
                                </table>
                                
                                <!-- Items Table -->
                                <table width="100%" cellpadding="0" cellspacing="0" style="margin-top: 20px; border: 1px solid #e5e7eb; border-radius: 8px; overflow: hidden;">
                                    <thead>
                                        <tr style="background-color: #f9fafb;">
                                            <th style="padding: 12px; text-align: left; color: #374151; font-weight: 600; font-size: 13px; text-transform: uppercase; letter-spacing: 0.5px;">Item</th>
                                            <th style="padding: 12px; text-align: center; color: #374151; font-weight: 600; font-size: 13px; text-transform: uppercase; letter-spacing: 0.5px;">Qty</th>
                                            <th style="padding: 12px; text-align: right; color: #374151; font-weight: 600; font-size: 13px; text-transform: uppercase; letter-spacing: 0.5px;">Price</th>
                                            <th style="padding: 12px; text-align: right; color: #374151; font-weight: 600; font-size: 13px; text-transform: uppercase; letter-spacing: 0.5px;">Total</th>
                                        </tr>
                                    </thead>
                                    <tbody>
                                        {items_html}
                                    </tbody>
                                </table>
                                
                                <!-- Totals -->
                                <table width="100%" cellpadding="0" cellspacing="0" style="margin-top: 20px;">
                                    <tr>
                                        <td style="padding: 8px 0; text-align: right; color: #6b7280;">Subtotal:</td>
                                        <td style="padding: 8px 0; text-align: right; width: 120px; color: #1f2937; font-weight: 500;">₹{b['subtotal']:.2f}</td>
                                    </tr>
                                    <tr>
                                        <td style="padding: 8px 0; text-align: right; color: #6b7280;">Discount:</td>
                                        <td style="padding: 8px 0; text-align: right; color: #dc2626; font-weight: 500;">- ₹{b['discount']:.2f}</td>
                                    </tr>
                                    <tr>
                                        <td style="padding: 15px 0 0 0; text-align: right; border-top: 2px solid #e5e7eb; color: #1f2937; font-size: 18px; font-weight: 600;">Total Amount:</td>
                                        <td style="padding: 15px 0 0 0; text-align: right; border-top: 2px solid #e5e7eb; color: #3b82f6; font-size: 20px; font-weight: 700;">₹{b['total_amount']:.2f}</td>
                                    </tr>
                                </table>
                            </td>
                        </tr>
                        
                        <!-- Footer -->
                        <tr>
                            <td style="background-color: #f9fafb; padding: 30px; text-align: center; border-top: 1px solid #e5e7eb;">
                                <p style="margin: 0 0 10px 0; color: #1f2937; font-size: 16px; font-weight: 500;">Thank you for your purchase!</p>
                                <p style="margin: 0; color: #6b7280; font-size: 13px;">For any queries, please contact us at the hospital pharmacy.</p>
                            </td>
                        </tr>
                    </table>
                    
                    <!-- Email Footer -->
                    <table width="600" cellpadding="0" cellspacing="0" style="margin-top: 20px;">
                        <tr>
                            <td style="text-align: center; color: #9ca3af; font-size: 12px; padding: 20px;">
                                <p style="margin: 0;">This is an automated email from Hospital Pharmacy System.</p>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """

@app.get("/api/download-backup")
def download_backup():
    """Download a zip file containing the database and Excel export of all data"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create in-memory zip file
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Add the database file
            if DB_PATH.exists():
                zip_file.write(DB_PATH, 'pharmacy.db')
            
            # Create Excel workbook with all data
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            conn = db()
            
            # Helper function to style headers
            def style_header(ws, row=1):
                for cell in ws[row]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Sheet 1: Medicines
            ws_medicines = wb.create_sheet("Medicines")
            medicines_data = conn.execute("""
                SELECT m.medicine_id, m.name, m.strength, m.category, m.manufacturer, 
                       m.reorder_level, m.is_active, m.created_at,
                       COALESCE(SUM(CASE WHEN b.expiry_date >= date('now') THEN b.quantity_available ELSE 0 END), 0) as total_stock
                FROM medicines m
                LEFT JOIN medicine_batches b ON b.medicine_id = m.medicine_id
                GROUP BY m.medicine_id
                ORDER BY m.name
            """).fetchall()
            
            ws_medicines.append(['Medicine ID', 'Name', 'Strength', 'Category', 'Manufacturer', 
                                'Reorder Level', 'Active', 'Created At', 'Total Stock'])
            style_header(ws_medicines)
            
            for row in medicines_data:
                ws_medicines.append([
                    row['medicine_id'], row['name'], row['strength'] or '', 
                    row['category'] or '', row['manufacturer'] or '',
                    row['reorder_level'], 'Yes' if row['is_active'] else 'No',
                    row['created_at'], row['total_stock']
                ])
            
            # Sheet 2: Batches
            ws_batches = wb.create_sheet("Medicine Batches")
            batches_data = conn.execute("""
                SELECT b.batch_id, m.name as medicine_name, b.batch_no, 
                       b.quantity_available, b.date_of_adding, b.expiry_date,
                       b.unit_price, b.supplier, b.created_at
                FROM medicine_batches b
                JOIN medicines m ON m.medicine_id = b.medicine_id
                ORDER BY b.created_at DESC
            """).fetchall()
            
            ws_batches.append(['Batch ID', 'Medicine', 'Batch No', 'Quantity Available', 
                              'Date Added', 'Expiry Date', 'Unit Price', 'Supplier', 'Created At'])
            style_header(ws_batches)
            
            for row in batches_data:
                ws_batches.append([
                    row['batch_id'], row['medicine_name'], row['batch_no'],
                    row['quantity_available'], row['date_of_adding'], row['expiry_date'],
                    row['unit_price'], row['supplier'] or '', row['created_at']
                ])
            
            # Sheet 3: Bills
            ws_bills = wb.create_sheet("Bills")
            bills_data = conn.execute("""
                SELECT b.bill_id, b.bill_no, b.customer_name, b.customer_email,
                       b.subtotal, b.discount, b.total_amount, b.payment_method,
                       u.name as created_by_name, b.created_at
                FROM bills b
                LEFT JOIN users u ON u.user_id = b.created_by
                ORDER BY b.created_at DESC
            """).fetchall()
            
            ws_bills.append(['Bill ID', 'Bill No', 'Customer Name', 'Customer Email',
                            'Subtotal', 'Discount', 'Total Amount', 'Payment Method',
                            'Created By', 'Created At'])
            style_header(ws_bills)
            
            for row in bills_data:
                ws_bills.append([
                    row['bill_id'], row['bill_no'], row['customer_name'] or 'Walk-in Customer',
                    row['customer_email'] or '', row['subtotal'], row['discount'],
                    row['total_amount'], row['payment_method'], row['created_by_name'] or '',
                    row['created_at']
                ])
            
            # Sheet 4: Bill Items
            ws_bill_items = wb.create_sheet("Bill Items")
            bill_items_data = conn.execute("""
                SELECT bi.bill_item_id, b.bill_no, m.name as medicine_name,
                       mb.batch_no, bi.quantity, bi.unit_price, bi.line_total
                FROM bill_items bi
                JOIN bills b ON b.bill_id = bi.bill_id
                JOIN medicines m ON m.medicine_id = bi.medicine_id
                JOIN medicine_batches mb ON mb.batch_id = bi.batch_id
                ORDER BY b.created_at DESC
            """).fetchall()
            
            ws_bill_items.append(['Item ID', 'Bill No', 'Medicine', 'Batch No',
                                 'Quantity', 'Unit Price', 'Line Total'])
            style_header(ws_bill_items)
            
            for row in bill_items_data:
                ws_bill_items.append([
                    row['bill_item_id'], row['bill_no'], row['medicine_name'],
                    row['batch_no'], row['quantity'], row['unit_price'], row['line_total']
                ])
            
            # Sheet 5: Stock Movements
            ws_movements = wb.create_sheet("Stock Movements")
            movements_data = conn.execute("""
                SELECT sm.movement_id, m.name as medicine_name, mb.batch_no,
                       sm.movement_type, sm.quantity_change, sm.reference_id,
                       sm.notes, u.name as created_by_name, sm.created_at
                FROM stock_movements sm
                JOIN medicines m ON m.medicine_id = sm.medicine_id
                LEFT JOIN medicine_batches mb ON mb.batch_id = sm.batch_id
                LEFT JOIN users u ON u.user_id = sm.created_by
                ORDER BY sm.created_at DESC
                LIMIT 500
            """).fetchall()
            
            ws_movements.append(['Movement ID', 'Medicine', 'Batch No', 'Movement Type',
                                'Quantity Change', 'Reference ID', 'Notes', 'Created By', 'Created At'])
            style_header(ws_movements)
            
            for row in movements_data:
                ws_movements.append([
                    row['movement_id'], row['medicine_name'], row['batch_no'] or '',
                    row['movement_type'], row['quantity_change'], row['reference_id'] or '',
                    row['notes'] or '', row['created_by_name'] or '', row['created_at']
                ])
            
            # Sheet 6: Users
            ws_users = wb.create_sheet("Users")
            users_data = conn.execute("""
                SELECT user_id, name, username, role, is_active, created_at
                FROM users
                ORDER BY role, name
            """).fetchall()
            
            ws_users.append(['User ID', 'Name', 'Username', 'Role', 'Active', 'Created At'])
            style_header(ws_users)
            
            for row in users_data:
                ws_users.append([
                    row['user_id'], row['name'], row['username'], row['role'],
                    'Yes' if row['is_active'] else 'No', row['created_at']
                ])
            
            conn.close()
            
            # Auto-adjust column widths for all sheets
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save Excel to memory
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Add Excel file to zip
            zip_file.writestr('pharmacy_data.xlsx', excel_buffer.getvalue())
        
        # Prepare zip for download
        zip_buffer.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'pharmacy_backup_{timestamp}.zip'
        
        return StreamingResponse(
            zip_buffer,
            media_type='application/zip',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
        
    except Exception as e:
        raise HTTPException(500, f"Failed to create backup: {str(e)}")

@app.post("/api/bills/{bill_id}/email")
def email_bill(bill_id:int, payload:EmailBillIn):
    conn=db(); b=get_bill_data(conn,bill_id); conn.close()
    smtp_server=payload.smtp_server or os.getenv("SMTP_SERVER","smtp.gmail.com")
    smtp_port=payload.smtp_port or int(os.getenv("SMTP_PORT","587"))
    smtp_username=payload.smtp_username or os.getenv("SMTP_USERNAME")
    smtp_password=payload.smtp_password or os.getenv("SMTP_PASSWORD")
    if not smtp_username or not smtp_password: raise HTTPException(400,"SMTP credentials missing. Set SMTP_USERNAME and SMTP_PASSWORD or enter them in the app.")
    msg=EmailMessage(); msg["Subject"]=f"Pharmacy Bill {b['bill_no']}"; msg["From"]=smtp_username; msg["To"]=payload.to_email
    # Set plain text version
    msg.set_content(bill_email_body_text(b))
    # Set HTML version (preferred by most email clients)
    msg.add_alternative(bill_email_body_html(b), subtype='html')
    try:
        with smtplib.SMTP(smtp_server,smtp_port) as server:
            server.starttls(); server.login(smtp_username,smtp_password); server.send_message(msg)
        return {"ok":True,"message":f"Bill shared with {payload.to_email}"}
    except Exception as e:
        raise HTTPException(500, f"Unable to send email: {e}")
